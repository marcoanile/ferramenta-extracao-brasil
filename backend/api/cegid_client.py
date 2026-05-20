"""Cegid Business integration stub.

Cegid does not expose a public REST API for external integrations.
The primary workflow is via the official Excel import templates.

This module provides helpers for generating Cegid-compatible import files.
When Cegid adds API support, extend this class with HTTP calls.
"""
import logging
from pathlib import Path

from parsers.banks.base import ParsedStatement

log = logging.getLogger(__name__)


class CegidClient:
    """Cegid Business integration — Excel template export."""

    def is_available(self) -> bool:
        return False  # No live API yet

    def generate_import_file(self, statement: ParsedStatement, output_path: Path,
                              client_name: str = "", year: int = None) -> Path:
        """Generate a Cegid Business compatible import Excel."""
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Border, Side

        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Movimentos"

        # Cegid standard import header
        headers = ["Data", "Diário", "Conta", "Subconta", "Descrição",
                   "Débito", "Crédito", "Nº Doc.", "Referência"]
        header_fill = PatternFill("solid", fgColor="003087")
        header_font = Font(bold=True, color="FFFFFF", name="Calibri", size=10)

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font

        col_widths = [12, 8, 12, 12, 40, 14, 14, 14, 20]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

        for row, mov in enumerate(statement.sorted_movements(), 2):
            debit = abs(mov.amount) if mov.amount < 0 else 0
            credit = mov.amount if mov.amount >= 0 else 0
            ws.cell(row=row, column=1, value=mov.date)
            ws.cell(row=row, column=2, value="")      # Diário — to fill manually
            ws.cell(row=row, column=3, value="")      # Conta — to fill manually
            ws.cell(row=row, column=4, value="")      # Subconta
            ws.cell(row=row, column=5, value=mov.description)
            ws.cell(row=row, column=6, value=debit or None).number_format = '#,##0.00'
            ws.cell(row=row, column=7, value=credit or None).number_format = '#,##0.00'
            ws.cell(row=row, column=8, value="")
            ws.cell(row=row, column=9, value=mov.reference)

        wb.save(output_path)
        log.info("Cegid import file saved: %s", output_path)
        return output_path


cegid = CegidClient()

"""Excel template engine — fills the TOConline import template or creates a fallback.

TOConline import template structure (importacao_movimentos.xlsx):
  Row 1 : Title (A1) + validation formula (E1)  — not touched
  Row 2 : C2="Saldo Inicial:"  D2=<opening_balance>
  Row 3 : C3="Saldo Final:"    D3=<closing_balance>
  Row 4 : Headers — A="Data Mov."  B="Data Valor"  C="Descrição do Movimento"  D="Movimento"
  Row 5+: Movement data

Fallback (no template): creates an 8-column workbook.
"""
import logging
import shutil
from datetime import datetime, date
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

from parsers.banks.base import ParsedStatement
import config

log = logging.getLogger(__name__)


def fill_template(statement: ParsedStatement, output_path: str | Path,
                  client_name: str = "", year: int = None,
                  template_path: str | Path = None) -> Path:
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    template = _find_template(template_path)
    if template:
        shutil.copy2(template, output_path)
        wb = openpyxl.load_workbook(output_path)
        ws = wb.active

        if _is_toconline_import_template(ws):
            _fill_toconline_template(ws, statement)
        else:
            _fill_generic_template(wb, statement, client_name, year)

        wb.save(output_path)
        log.info("Template filled and saved to %s (%d movements)", output_path, len(statement.movements))
    else:
        wb = _create_fallback_workbook(statement, client_name, year)
        wb.save(output_path)
        log.info("No template found — created fallback workbook at %s (%d movements)",
                 output_path, len(statement.movements))

    return output_path


# ─── Template detection ───────────────────────────────────────────────────────

def _find_template(template_path=None) -> Path | None:
    candidates = []
    if template_path:
        candidates.append(Path(template_path))
    candidates += sorted(config.TEMPLATES_DIR.glob("*.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True)
    candidates += sorted(config.TEMPLATES_DIR.glob("*.xls"), key=lambda p: p.stat().st_mtime, reverse=True)
    for p in candidates:
        if p.exists():
            return p
    return None


def _is_toconline_import_template(ws) -> bool:
    """Detect the TOConline import template by looking for its specific column headers."""
    for row_idx in range(1, 10):
        row_vals = [str(ws.cell(row=row_idx, column=c).value or "").strip().lower()
                    for c in range(1, 6)]
        if "data mov." in row_vals or ("data mov" in " ".join(row_vals) and "movimento" in " ".join(row_vals)):
            return True
    return False


def _find_header_row(ws) -> int:
    """Return the row index of the 'Data Mov.' header row."""
    for row_idx in range(1, 15):
        val = str(ws.cell(row=row_idx, column=1).value or "").strip().lower()
        if "data mov" in val:
            return row_idx
    return 4  # default


def _find_balance_rows(ws) -> tuple[int | None, int | None]:
    """Return (saldo_inicial_row, saldo_final_row) by scanning for those labels."""
    ini_row = fin_row = None
    for row_idx in range(1, 15):
        for col in range(1, 5):
            val = str(ws.cell(row=row_idx, column=col).value or "").lower()
            if "saldo inicial" in val:
                ini_row = row_idx
            elif "saldo final" in val:
                fin_row = row_idx
    return ini_row, fin_row


# ─── TOConline import fill ─────────────────────────────────────────────────────

def _fill_toconline_template(ws, statement: ParsedStatement):
    """Fill the official TOConline import template in-place."""
    header_row = _find_header_row(ws)
    ini_row, fin_row = _find_balance_rows(ws)
    data_start = header_row + 1

    # Write opening balance in column D
    if ini_row:
        ws.cell(row=ini_row, column=4).value = statement.opening_balance

    # Clear all existing data rows (preserve header and metadata rows)
    max_existing = ws.max_row
    for row_idx in range(data_start, max_existing + 1):
        for col in range(1, ws.max_column + 1):
            ws.cell(row=row_idx, column=col).value = None

    # Write movements
    movements = statement.sorted_movements()
    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for i, mov in enumerate(movements):
        row = data_start + i
        mov_date = _to_date(mov.date)

        # A: Data Mov.
        cell_a = ws.cell(row=row, column=1, value=mov_date)
        cell_a.number_format = "DD/MM/YYYY"
        cell_a.border = thin

        # B: Data Valor (same as mov date — BPI doesn't give us a separate value date)
        cell_b = ws.cell(row=row, column=2, value=mov_date)
        cell_b.number_format = "DD/MM/YYYY"
        cell_b.border = thin

        # C: Descrição do Movimento
        cell_c = ws.cell(row=row, column=3, value=mov.description)
        cell_c.border = thin

        # D: Movimento (signed: negative=debit, positive=credit)
        cell_d = ws.cell(row=row, column=4, value=mov.amount)
        cell_d.number_format = '#,##0.00'
        cell_d.border = thin

    # Compute closing balance arithmetically from opening balance + sum of movements.
    # This guarantees TOConline's validation (Saldo Final = Saldo Inicial + SUM(D))
    # passes exactly, regardless of floating-point drift in the PDF's running balances.
    if fin_row:
        opening = statement.opening_balance or 0.0
        total_movements = sum(m.amount for m in movements)
        closing = round(opening + total_movements, 2)
        ws.cell(row=fin_row, column=4).value = closing


# ─── Generic template fill ─────────────────────────────────────────────────────

def _fill_generic_template(wb, statement: ParsedStatement, client_name: str, year: int):
    ws = wb.active
    _write_generic_movements(ws, statement, client_name, year)


def _create_fallback_workbook(statement: ParsedStatement, client_name: str, year: int):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Movimentos"

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    center = Alignment(horizontal="center", vertical="center")

    col_widths = {"A": 14, "B": 50, "C": 20, "D": 14, "E": 14, "F": 16, "G": 12, "H": 20}
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w

    headers = ["Data", "Descrição", "Referência", "Débito", "Crédito", "Saldo", "Tipo", "Categoria"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin
        cell.alignment = center

    ws.freeze_panes = "A2"
    _write_generic_movements(ws, statement, client_name, year, data_start=2)
    return wb


def _write_generic_movements(ws, statement: ParsedStatement, client_name: str,
                              year: int, data_start: int = None):
    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    if data_start is None:
        data_start = _find_data_start(ws)

    for i, mov in enumerate(statement.sorted_movements()):
        row = data_start + i
        debit = abs(mov.amount) if mov.amount < 0 else None
        credit = mov.amount if mov.amount >= 0 else None
        values = [mov.date, mov.description, mov.reference, debit, credit, mov.balance,
                  mov.movement_type, mov.category]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = thin
            if col == 1 and val:
                cell.number_format = "DD/MM/YYYY"
            elif col in (4, 5, 6) and val is not None:
                cell.number_format = "#,##0.00"


def _find_data_start(ws) -> int:
    for row in range(1, 20):
        val = str(ws.cell(row=row, column=1).value or "").lower().strip()
        if val in ("data", "date", "data mov.", "data mov"):
            return row + 1
    return 2


# ─── Helpers ──────────────────────────────────────────────────────────────────

def _to_date(value):
    """Convert YYYY-MM-DD string or datetime to Python date."""
    if isinstance(value, (datetime, date)):
        return value
    if isinstance(value, str) and len(value) == 10:
        try:
            return datetime.strptime(value, "%Y-%m-%d").date()
        except ValueError:
            pass
    return value

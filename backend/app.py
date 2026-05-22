"""Nexus Extrator — Flask backend."""
import io
import logging
import sys
import uuid
from datetime import datetime
from pathlib import Path

from flask import Flask, request, jsonify, send_file
from flask_cors import CORS

import config
from storage.database import (
    init_db, get_groups, get_group, create_group, delete_group,
    get_extracts, add_extract, delete_extract, add_movements,
    get_movements_for_group,
)
from parsers.pdf_parser import parse_pdf
from parsers.excel_parser import parse_excel
from parsers.csv_parser import parse_csv
from converter.template_engine import (
    fill_template, _find_template, _find_header_row, _find_balance_rows, _to_date,
)

logging.basicConfig(
    level=logging.DEBUG if config.DEBUG else logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
    handlers=[
        logging.StreamHandler(sys.stdout),
        logging.FileHandler(config.LOGS_DIR / "nexus.log", encoding="utf-8"),
    ],
)
log = logging.getLogger("nexus")

app = Flask(__name__, static_folder="../frontend/static", static_url_path="/static")
app.secret_key = config.SECRET_KEY
CORS(app, origins="*")

init_db()


@app.route("/")
def index():
    return send_file("../frontend/index.html")


def _parse_file(file_path: Path, suffix: str):
    if suffix == ".pdf":
        return parse_pdf(file_path, "")
    if suffix in (".xlsx", ".xls"):
        return parse_excel(file_path, "")
    if suffix == ".csv":
        return parse_csv(file_path, "")
    raise ValueError(f"Formato não suportado: {suffix}")


# ── Stateless converter ───────────────────────────────────────────────────────

@app.route("/convert", methods=["POST"])
def convert():
    if "file" not in request.files:
        return jsonify({"error": "Ficheiro em falta"}), 400
    file = request.files["file"]
    suffix = Path(file.filename).suffix.lower()
    if suffix not in config.SUPPORTED_FORMATS:
        return jsonify({"error": f"Formato não suportado: {suffix}"}), 400

    tmp = config.TEMP_DIR / f"{uuid.uuid4().hex}{suffix}"
    out = config.TEMP_DIR / f"extrato_{uuid.uuid4().hex[:8]}.xlsx"
    file.save(str(tmp))

    try:
        parsed = _parse_file(tmp, suffix)
        movements = parsed.sorted_movements()
        year = int(parsed.period_start[:4]) if parsed.period_start else datetime.now().year
        fill_template(statement=parsed, output_path=out, client_name="", year=year)

        bank_slug = (parsed.bank_name or "extrato").lower().replace(" ", "_").replace("/", "-")
        download_name = f"extrato_{bank_slug}_{year}.xlsx"

        buf = io.BytesIO(out.read_bytes())
        buf.seek(0)
        tmp.unlink(missing_ok=True)
        out.unlink(missing_ok=True)

        response = send_file(buf, as_attachment=True, download_name=download_name,
                             mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response.headers["X-Bank-Name"] = parsed.bank_name or ""
        response.headers["X-Movement-Count"] = str(len(movements))
        response.headers["X-Period-Start"] = parsed.period_start or ""
        response.headers["X-Period-End"] = parsed.period_end or ""
        return response

    except Exception as e:
        log.exception("Conversion error")
        tmp.unlink(missing_ok=True)
        out.unlink(missing_ok=True)
        return jsonify({"error": str(e)}), 500


# ── Consolidation groups ──────────────────────────────────────────────────────

@app.route("/api/groups", methods=["GET"])
def list_groups():
    return jsonify({"data": get_groups()})


@app.route("/api/groups", methods=["POST"])
def create_group_route():
    body = request.get_json() or {}
    name = (body.get("name") or "").strip()
    if not name:
        return jsonify({"error": "Nome obrigatório"}), 400
    gid = create_group(name)
    return jsonify({"id": gid, "name": name}), 201


@app.route("/api/groups/<int:group_id>", methods=["DELETE"])
def delete_group_route(group_id: int):
    if not get_group(group_id):
        return jsonify({"error": "Grupo não encontrado"}), 404
    delete_group(group_id)
    return jsonify({"ok": True})


@app.route("/api/groups/<int:group_id>/extracts", methods=["GET"])
def list_extracts(group_id: int):
    if not get_group(group_id):
        return jsonify({"error": "Grupo não encontrado"}), 404
    return jsonify({"data": get_extracts(group_id)})


@app.route("/api/groups/<int:group_id>/upload", methods=["POST"])
def upload_to_group(group_id: int):
    try:
        group = get_group(group_id)
    except Exception as e:
        log.exception("upload_to_group DB error")
        return jsonify({"error": str(e)}), 500
    if not group:
        return jsonify({"error": "Grupo não encontrado"}), 404
    if "file" not in request.files:
        return jsonify({"error": "Ficheiro em falta"}), 400
    file = request.files["file"]
    suffix = Path(file.filename).suffix.lower()
    if suffix not in config.SUPPORTED_FORMATS:
        return jsonify({"error": f"Formato não suportado: {suffix}"}), 400

    tmp = config.TEMP_DIR / f"{uuid.uuid4().hex}{suffix}"
    file.save(str(tmp))

    try:
        parsed = _parse_file(tmp, suffix)
        movements = parsed.sorted_movements()
        extract_id = add_extract(
            group_id=group_id,
            filename=file.filename,
            bank_name=parsed.bank_name,
            period_start=parsed.period_start,
            period_end=parsed.period_end,
            movement_count=len(movements),
        )
        add_movements(extract_id, group_id, [
            {"date": m.date, "description": m.description,
             "amount": m.amount, "balance": m.balance}
            for m in movements
        ])
        return jsonify({
            "extract_id": extract_id,
            "bank_name": parsed.bank_name,
            "period_start": parsed.period_start,
            "period_end": parsed.period_end,
            "movement_count": len(movements),
        }), 201

    except Exception as e:
        log.exception("Upload-to-group error")
        return jsonify({"error": str(e)}), 500
    finally:
        tmp.unlink(missing_ok=True)


@app.route("/api/groups/<int:group_id>/extracts/<int:extract_id>", methods=["DELETE"])
def delete_extract_route(group_id: int, extract_id: int):
    delete_extract(extract_id)
    return jsonify({"ok": True})


@app.route("/api/groups/<int:group_id>/generate")
def generate_group(group_id: int):
    try:
        group = get_group(group_id)
        movements = get_movements_for_group(group_id)
    except Exception as e:
        log.exception("generate_group DB error")
        return jsonify({"error": str(e)}), 500
    if not group:
        return jsonify({"error": "Grupo não encontrado"}), 404
    if not movements:
        return jsonify({"error": "Sem movimentos para consolidar"}), 400

    out = config.TEMP_DIR / f"consolidado_{uuid.uuid4().hex[:8]}.xlsx"
    try:
        _build_consolidated_excel(movements, out)
        safe_name = group["name"].replace(" ", "_").replace("/", "-")

        buf = io.BytesIO(out.read_bytes())
        buf.seek(0)
        out.unlink(missing_ok=True)

        return send_file(buf, as_attachment=True, download_name=f"{safe_name}.xlsx",
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    except Exception as e:
        log.exception("Generate error")
        out.unlink(missing_ok=True)
        return jsonify({"error": str(e)}), 500


def _build_consolidated_excel(movements: list[dict], out: Path):
    import shutil
    import openpyxl
    from openpyxl.styles import Border, Side

    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    template = _find_template(None)

    if template:
        shutil.copy2(str(template), str(out))
        wb = openpyxl.load_workbook(str(out))
        ws = wb.active
        header_row = _find_header_row(ws)
        ini_row, fin_row = _find_balance_rows(ws)
        data_start = header_row + 1

        first_m = next((m for m in movements if m["balance"] is not None), None)
        last_m = next((m for m in reversed(movements) if m["balance"] is not None), None)
        opening = round((first_m["balance"] or 0) - (first_m["amount"] or 0), 2) if first_m else 0
        effective_closing = (last_m["balance"] or 0) if last_m else 0

        if ini_row:
            ws.cell(row=ini_row, column=4).value = opening
        if fin_row:
            ws.cell(row=fin_row, column=4).value = effective_closing

        for row_idx in range(data_start, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row_idx, column=col).value = None

        for i, mov in enumerate(movements):
            row = data_start + i
            d = _to_date(mov["date"])
            ws.cell(row=row, column=1, value=d).number_format = "DD/MM/YYYY"
            ws.cell(row=row, column=1).border = thin
            ws.cell(row=row, column=2, value=d).number_format = "DD/MM/YYYY"
            ws.cell(row=row, column=2).border = thin
            ws.cell(row=row, column=3, value=mov.get("description") or "").border = thin
            cell = ws.cell(row=row, column=4, value=mov["amount"])
            cell.number_format = "#,##0.00"
            cell.border = thin
        wb.save(str(out))
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Movimentos"
        ws.append(["Data Mov.", "Data Valor", "Descrição do Movimento", "Movimento"])
        for mov in movements:
            d = _to_date(mov["date"])
            ws.append([d, d, mov.get("description") or "", mov["amount"]])
        wb.save(str(out))


@app.route("/api/health")
def health():
    return jsonify({"status": "ok"})


if __name__ == "__main__":
    log.info("Starting Nexus on port %d", config.PORT)
    app.run(host="127.0.0.1", port=config.PORT, debug=config.DEBUG)
